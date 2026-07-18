import csv
import io
import math
import random
import re
import struct
import wave
from datetime import datetime, date
from fastapi import APIRouter, Depends, HTTPException, Request, Response, UploadFile, File, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from typing import Optional, List
from app.api.deps import get_db, get_current_user, check_roles
from app.models.models import PipelineLead, ImportBatch, Tenant
from app.schemas.schemas import (
    PipelineLeadCreate, PipelineLeadUpdate, PipelineLeadResponse, PipelineLeadPage,
    PipelineStats, BulkMoveRequest, BulkMoveResult, ImportBatchResponse,
    LeadSettingResponse, LeadSettingUpdate, AICallResult
)
from app.services import pipeline_service, ai_agent

router = APIRouter()

ADMIN_ROLES = ["Super Admin", "Tenant Admin"]

SORTABLE_FIELDS = {
    "id", "date", "name", "phone", "email", "source", "project", "budget",
    "status", "interest_status", "called_at", "call_duration", "contacted_by",
    "next_followup_date", "site_visit_status"
}

IMPORT_COLUMNS = {"date", "name", "phone", "email", "source", "project", "budget", "status"}


async def tenant_name(request: Request, db: AsyncSession) -> str:
    tenant_id = getattr(request.state, "tenant_id", "public")
    result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    tenant = result.scalars().first()
    return tenant.name if tenant else tenant_id


# ========================================================
# STATS (nav badges)
# ========================================================

@router.get("/stats", response_model=PipelineStats)
async def pipeline_stats(db: AsyncSession = Depends(get_db), user=Depends(get_current_user)):
    counts = {}
    for stage in ("raw", "called", "qualified"):
        result = await db.execute(
            select(func.count()).select_from(PipelineLead).where(PipelineLead.stage == stage)
        )
        counts[stage] = result.scalar() or 0
    return counts


# ========================================================
# LIST / CRUD
# ========================================================

@router.get("/leads", response_model=PipelineLeadPage)
async def list_pipeline_leads(
    stage: str = "raw",
    search: Optional[str] = None,
    source: Optional[str] = None,
    project: Optional[str] = None,
    lead_status: Optional[str] = None,
    interest: Optional[str] = None,
    sort: str = "date",
    order: str = "desc",
    page: int = 1,
    limit: int = 10,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user)
):
    """Paginated, filterable, sortable listing for any pipeline stage."""
    result = await db.execute(select(PipelineLead).where(PipelineLead.stage == stage))
    leads = result.scalars().all()

    filtered = []
    for lead in leads:
        if search:
            hay = f"{lead.id} {lead.name} {lead.phone} {lead.email} {lead.project}".lower()
            if search.lower() not in hay:
                continue
        if source and lead.source != source:
            continue
        if project and lead.project != project:
            continue
        if lead_status and lead.status != lead_status:
            continue
        if interest and (lead.interest_status or "") != interest:
            continue
        filtered.append(lead)

    sort_key = sort if sort in SORTABLE_FIELDS else "date"
    reverse = order != "asc"
    filtered.sort(key=lambda l: str(getattr(l, sort_key, "") or ""), reverse=reverse)

    total = len(filtered)
    pages = max(math.ceil(total / limit), 1)
    page = min(max(page, 1), pages)
    items = filtered[(page - 1) * limit: page * limit]
    return {"items": items, "total": total, "page": page, "pages": pages}


@router.get("/leads/{lead_id}", response_model=PipelineLeadResponse)
async def get_pipeline_lead(lead_id: str, db: AsyncSession = Depends(get_db), user=Depends(get_current_user)):
    result = await db.execute(select(PipelineLead).where(PipelineLead.id == lead_id))
    lead = result.scalars().first()
    if not lead:
        raise HTTPException(status_code=404, detail="Pipeline lead not found")
    return lead


# Manual-entry defaults per target module: (status, stage label)
MANUAL_ENTRY_STAGES = {
    "raw": ("Raw Lead", "Raw Leads"),
    "called": ("Called Lead", "Called Leads"),
    "qualified": ("Qualified Lead", "Qualified Leads"),
}


@router.post("/leads", response_model=PipelineLeadResponse, status_code=status.HTTP_201_CREATED)
async def create_pipeline_lead(
    payload: PipelineLeadCreate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user)
):
    """
    Registers an individual lead directly into the Raw, Called, or Qualified
    module. All entry points share the same duplicate validation, history,
    and audit logging.
    """
    target = payload.stage or "raw"
    if target not in MANUAL_ENTRY_STAGES:
        raise HTTPException(status_code=400, detail=f"Leads cannot be created directly in stage '{target}'")
    entry_status, stage_label = MANUAL_ENTRY_STAGES[target]

    settings_row = await pipeline_service.get_settings(db)
    clash = await pipeline_service.find_duplicate(db, payload.phone, payload.email, settings_row)
    if clash:
        raise HTTPException(status_code=409, detail=f"Duplicate lead rejected: {clash}")

    new_id = await pipeline_service.next_pipeline_id(db)
    lead = PipelineLead(
        id=new_id,
        date=datetime.now().strftime("%Y-%m-%d"),
        name=payload.name,
        phone=payload.phone,
        email=payload.email,
        source=payload.source,
        project=payload.project,
        budget=payload.budget,
        stage=target,
        status=entry_status,
        call_attempts=0,
        history=[{"date": pipeline_service.now_stamp(),
                  "action": f"Lead registered manually in {stage_label}", "user": user.username}]
    )

    if target == "called":
        lead.interest_status = payload.interest_status or "Interested"
        lead.called_at = pipeline_service.now_stamp()
        lead.ai_outcome = "Manual call entry by sales user"
    elif target == "qualified":
        lead.contacted_by = payload.contacted_by or user.username
        lead.remarks = payload.remarks
        lead.site_visit_status = payload.site_visit_status or "Not Scheduled"
        lead.loan_requirement = payload.loan_requirement or "Pending Assessment"
        lead.next_followup_date = payload.next_followup_date

    db.add(lead)
    await pipeline_service.write_audit(db, await tenant_name(request, db), user.username,
                                       f"Lead {new_id} ({payload.name}) created manually in {stage_label}.")
    await db.commit()
    return lead


@router.put("/leads/{lead_id}", response_model=PipelineLeadResponse)
async def update_pipeline_lead(
    lead_id: str,
    payload: PipelineLeadUpdate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user)
):
    result = await db.execute(select(PipelineLead).where(PipelineLead.id == lead_id))
    lead = result.scalars().first()
    if not lead:
        raise HTTPException(status_code=404, detail="Pipeline lead not found")

    update_data = payload.dict(exclude_unset=True)
    changed = [k for k, v in update_data.items() if getattr(lead, k) != v]
    for key, value in update_data.items():
        setattr(lead, key, value)

    if changed:
        pipeline_service.add_history(lead, f"Details updated: {', '.join(changed)}", user.username)
        await pipeline_service.write_audit(db, await tenant_name(request, db), user.username,
                                           f"Pipeline lead {lead_id} updated ({', '.join(changed)}).")
    await db.commit()
    return lead


@router.delete("/leads/{lead_id}")
async def delete_pipeline_lead(
    lead_id: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user=Depends(check_roles(ADMIN_ROLES))
):
    result = await db.execute(select(PipelineLead).where(PipelineLead.id == lead_id))
    lead = result.scalars().first()
    if not lead:
        raise HTTPException(status_code=404, detail="Pipeline lead not found")
    await db.delete(lead)
    await pipeline_service.write_audit(db, await tenant_name(request, db), user.username,
                                       f"Pipeline lead {lead_id} ({lead.name}) deleted.")
    await db.commit()
    return {"deleted": lead_id}


# ========================================================
# STAGE MOVEMENT (single + bulk, transaction-safe)
# ========================================================

@router.post("/leads/bulk-move", response_model=BulkMoveResult)
async def bulk_move_leads(
    payload: BulkMoveRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user)
):
    if payload.target not in {"called", "qualified", "rejected", "database", "customer"}:
        raise HTTPException(status_code=400, detail="Unknown target stage")

    tenant = await tenant_name(request, db)
    moved, skipped, notes = 0, 0, []
    for lead_id in payload.ids:
        result = await db.execute(select(PipelineLead).where(PipelineLead.id == lead_id))
        lead = result.scalars().first()
        if not lead:
            skipped += 1
            notes.append(f"{lead_id}: not found")
            continue
        error = await pipeline_service.move_lead(db, lead, payload.target, user.username, tenant)
        if error:
            skipped += 1
            notes.append(error)
        else:
            moved += 1

    await db.commit()
    label = pipeline_service.STAGE_LABELS.get(payload.target, payload.target)
    detail = f"Moved {moved} lead(s) to {label}." + (f" Skipped: {'; '.join(notes)}" if notes else "")
    return {"moved": moved, "skipped": skipped, "detail": detail}


# ========================================================
# BULK IMPORT / EXPORT / IMPORT HISTORY
# ========================================================

def _cell_to_text(value) -> str:
    """
    Normalizes an Excel cell to clean text. Integral floats (how Excel stores
    typed-in numbers like phone digits) become plain digit strings instead of
    '9876543210.0' or scientific notation; date cells become YYYY-MM-DD.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _parse_rows(filename: str, content: bytes) -> List[dict]:
    """Parses CSV or Excel uploads into normalized dict rows (shared workflow)."""
    rows: List[dict] = []
    if filename.lower().endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = []
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx == 0:
                headers = [str(h or "").strip().lower().replace(" ", "_") for h in row]
                continue
            parsed = {headers[i]: _cell_to_text(v) for i, v in enumerate(row) if i < len(headers)}
            if any(v for v in parsed.values()):  # skip fully blank spreadsheet rows
                rows.append(parsed)
        wb.close()
    else:
        text_content = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text_content))
        for row in reader:
            parsed = {(k or "").strip().lower().replace(" ", "_"): (v or "").strip() for k, v in row.items()}
            if any(v for v in parsed.values()):
                rows.append(parsed)
    return rows


EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _validate_import_row(row: dict) -> tuple[dict | None, str | None]:
    """
    Bulk-import validation. The Mobile Number is the ONLY mandatory field:
    a row imports as long as it carries a valid mobile number, even if every
    other field is blank. All other fields (name, email, source, project,
    budget) are optional and can be filled in later through the CRM.
    Returns (normalized_fields, None) on success or (None, reason) on failure.
    """
    name = (row.get("name") or "").strip()
    phone = str(row.get("phone") or row.get("phone_number") or row.get("mobile") or "").strip()
    email = (row.get("email") or row.get("email_id") or "").strip()

    # Mobile Number — the single required, validated field
    if not phone:
        return None, "missing mobile number"
    if len(pipeline_service.normalize_phone(phone)) < 10:
        return None, f"invalid mobile number '{phone}' (needs at least 10 digits)"

    return {"name": name, "phone": phone, "email": email}, None


@router.post("/import")
async def bulk_import_leads(
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user)
):
    """
    Bulk lead upload from Excel/CSV. Headers (case-insensitive): name, phone,
    email, source, project, and optionally date, budget, status. Every row is
    duplicate-validated per the workspace policy; the batch is recorded in
    import history and the audit trail.
    """
    content = await file.read()
    try:
        rows = _parse_rows(file.filename or "upload.csv", content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {exc}")

    if not rows:
        raise HTTPException(status_code=400, detail="File contained no data rows")

    settings_row = await pipeline_service.get_settings(db)
    imported, duplicates, errors = 0, 0, 0
    error_details = []
    duplicate_details = []

    for idx, row in enumerate(rows, start=2):
        fields, problem = _validate_import_row(row)
        if problem:
            errors += 1
            error_details.append(f"Row {idx}: {problem}")
            continue

        clash = await pipeline_service.find_duplicate(db, fields["phone"], fields["email"], settings_row)
        if clash:
            duplicates += 1
            duplicate_details.append(f"Row {idx} ({fields['name'] or fields['phone']}): {clash}")
            continue

        new_id = await pipeline_service.next_pipeline_id(db)
        # Optional fields fall back to safe placeholders so a mobile-only row
        # still satisfies the NOT NULL columns; users refine these later.
        lead = PipelineLead(
            id=new_id,
            date=row.get("date") or datetime.now().strftime("%Y-%m-%d"),
            name=fields["name"] or "Unnamed Lead",
            phone=fields["phone"],
            email=fields["email"] or f"unknown+{new_id.lower()}@leads.import",
            source=row.get("source") or "Bulk Import",
            project=row.get("project") or row.get("project_name") or "Unassigned",
            budget=row.get("budget") or None,
            stage="raw",
            status="Pending Call",
            call_attempts=0,
            history=[{"date": pipeline_service.now_stamp(),
                      "action": f"Imported via bulk upload ({file.filename})", "user": user.username}]
        )
        db.add(lead)
        await db.flush()
        imported += 1

    batch = ImportBatch(
        filename=file.filename or "upload.csv",
        total_rows=len(rows),
        imported=imported,
        duplicates=duplicates,
        errors=errors,
        uploaded_by=user.username
    )
    db.add(batch)
    await pipeline_service.write_audit(
        db, await tenant_name(request, db), user.username,
        f"Bulk import '{file.filename}': {imported} imported, {duplicates} duplicates blocked, {errors} errors."
    )
    await db.commit()
    return {
        "filename": file.filename, "total_rows": len(rows), "imported": imported,
        "duplicates": duplicates, "errors": errors,
        "error_details": error_details[:20], "duplicate_details": duplicate_details[:20]
    }


# Identical column order and sample data for BOTH template formats
TEMPLATE_COLUMNS = ["name", "phone", "email", "source", "project", "budget"]
TEMPLATE_SAMPLE_ROWS = [
    ["Sample Lead", "9876543210", "sample@gmail.com", "Website Form", "Sunrise Heights", "₹90 Lakhs"],
    ["Asha Verma", "09123456789", "asha.verma@yahoo.com", "MagicBricks", "Green Meadows", "₹1.2 Crore"],
]


@router.get("/import-template")
async def download_import_template(
    format: str = "csv",
    user=Depends(get_current_user)
):
    """
    Serves the bulk-import template. Both formats share the same columns,
    order, and sample data; the Excel variant formats the phone column as
    text so leading zeros survive and long numbers never collapse into
    scientific notation.
    """
    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Raw Leads Import"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        for col_idx, header in enumerate(TEMPLATE_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

        phone_col = TEMPLATE_COLUMNS.index("phone") + 1
        for row_idx, sample in enumerate(TEMPLATE_SAMPLE_ROWS, start=2):
            for col_idx, value in enumerate(sample, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == phone_col:
                    cell.number_format = "@"

        # Pre-format the phone column as text for the next ~500 data rows so
        # values typed or pasted by the user keep leading zeros as well.
        for row_idx in range(2, 502):
            ws.cell(row=row_idx, column=phone_col).number_format = "@"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=raw_leads_template.xlsx"}
        )

    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(TEMPLATE_COLUMNS)
    for sample in TEMPLATE_SAMPLE_ROWS:
        writer.writerow(sample)
    return Response(
        content=out.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=raw_leads_template.csv"}
    )


@router.get("/imports", response_model=List[ImportBatchResponse])
async def import_history(db: AsyncSession = Depends(get_db), user=Depends(get_current_user)):
    result = await db.execute(select(ImportBatch).order_by(ImportBatch.date.desc()))
    return result.scalars().all()


EXPORT_FIELDS = {
    "raw": ["id", "date", "name", "phone", "email", "source", "project", "budget", "status", "call_attempts"],
    "called": ["id", "name", "phone", "email", "source", "project", "budget", "interest_status",
               "called_at", "call_duration", "ai_outcome", "ai_confidence"],
    "qualified": ["id", "name", "phone", "email", "source", "project", "budget", "contacted_by",
                  "remarks", "site_visit_status", "loan_requirement", "next_followup_date"],
}


@router.get("/export")
async def export_leads(
    request: Request,
    stage: str = "raw",
    format: str = "csv",
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user)
):
    """Exports a stage's leads as CSV or Excel."""
    fields = EXPORT_FIELDS.get(stage, EXPORT_FIELDS["raw"])
    result = await db.execute(select(PipelineLead).where(PipelineLead.stage == stage))
    leads = result.scalars().all()

    await pipeline_service.write_audit(db, await tenant_name(request, db), user.username,
                                       f"Exported {len(leads)} {stage} lead(s) as {format.upper()}.")
    await db.commit()

    if format == "xlsx":
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{stage.capitalize()} Leads"
        ws.append([f.replace("_", " ").title() for f in fields])
        for lead in leads:
            ws.append([str(getattr(lead, f, "") or "") for f in fields])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={stage}_leads.xlsx"}
        )

    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow([f.replace("_", " ").title() for f in fields])
    for lead in leads:
        writer.writerow([str(getattr(lead, f, "") or "") for f in fields])
    return Response(
        content=out.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={stage}_leads.csv"}
    )


# ========================================================
# CALL RECORDING (synthesized placeholder audio)
# ========================================================

@router.get("/leads/{lead_id}/recording")
async def get_call_recording(
    lead_id: str,
    request: Request,
    download: bool = False,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user)
):
    """
    Streams the AI call recording. Playback is open to all roles; downloading
    the file is an admin-level permission.
    """
    if download and user.role not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Downloading recordings requires an admin role")

    result = await db.execute(select(PipelineLead).where(PipelineLead.id == lead_id))
    lead = result.scalars().first()
    if not lead:
        raise HTTPException(status_code=404, detail="Pipeline lead not found")
    if not lead.recording_available:
        raise HTTPException(status_code=404, detail="No call recording exists for this lead")

    # Deterministic synthesized "conversation" tones seeded by the lead id, so
    # every lead has a stable, unique placeholder recording without file storage.
    seed = sum(ord(c) for c in lead_id)
    rng = random.Random(seed)
    sample_rate = 8000
    buf = io.BytesIO()
    with wave.open(buf, "wb") as wav:
        wav.setnchannels(1)
        wav.setsampwidth(2)
        wav.setframerate(sample_rate)
        frames = bytearray()
        for _segment in range(rng.randint(6, 10)):
            freq = rng.choice([220, 260, 300, 340, 400, 460])
            seg_len = rng.uniform(0.25, 0.7)
            volume = rng.uniform(0.2, 0.5)
            for i in range(int(seg_len * sample_rate)):
                t = i / sample_rate
                # two-tone blend with a light tremolo reads as "voice-like" enough for a demo
                value = volume * (math.sin(2 * math.pi * freq * t) * 0.7 +
                                  math.sin(2 * math.pi * (freq * 1.5) * t) * 0.3)
                value *= 0.75 + 0.25 * math.sin(2 * math.pi * 6 * t)
                frames += struct.pack("<h", int(value * 32767))
            frames += b"\x00\x00" * int(0.12 * sample_rate)
        wav.writeframes(bytes(frames))
    buf.seek(0)

    if download:
        await pipeline_service.write_audit(db, await tenant_name(request, db), user.username,
                                           f"Call recording for {lead_id} downloaded.")
        await db.commit()

    headers = {"Content-Disposition": f"{'attachment' if download else 'inline'}; filename={lead_id}_recording.wav"}
    return StreamingResponse(buf, media_type="audio/wav", headers=headers)


# ========================================================
# SETTINGS (duplicate policy + AI calling config)
# ========================================================

@router.get("/settings", response_model=LeadSettingResponse)
async def get_pipeline_settings(db: AsyncSession = Depends(get_db), user=Depends(get_current_user)):
    settings_row = await pipeline_service.get_settings(db)
    await db.commit()
    return settings_row


@router.put("/settings", response_model=LeadSettingResponse)
async def update_pipeline_settings(
    payload: LeadSettingUpdate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user=Depends(check_roles(ADMIN_ROLES))
):
    settings_row = await pipeline_service.get_settings(db)
    update_data = payload.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(settings_row, key, value)
    await pipeline_service.write_audit(db, await tenant_name(request, db), user.username,
                                       f"Pipeline settings updated: {', '.join(update_data)}.")
    await db.commit()
    return settings_row


# ========================================================
# AI AGENT INTEGRATION (REST surface for external dialers + demo trigger)
# ========================================================

@router.get("/ai/pending", response_model=List[PipelineLeadResponse])
async def ai_pending_leads(limit: int = 10, db: AsyncSession = Depends(get_db), user=Depends(get_current_user)):
    """Returns raw leads awaiting an AI call — the pickup queue for external dialer integrations."""
    settings_row = await pipeline_service.get_settings(db)
    result = await db.execute(
        select(PipelineLead)
        .where(PipelineLead.stage == "raw")
        .where(PipelineLead.status.in_(list(ai_agent.PENDING_STATUSES)))
        .where(PipelineLead.call_attempts < settings_row.ai_retry_limit)
        .order_by(PipelineLead.created_at)
        .limit(limit)
    )
    await db.commit()
    return result.scalars().all()


@router.post("/ai/call-result", response_model=PipelineLeadResponse)
async def ai_call_result(
    payload: AICallResult,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user)
):
    """Webhook-style endpoint where an AI dialer posts a completed/failed call result."""
    result = await db.execute(select(PipelineLead).where(PipelineLead.id == payload.lead_id))
    lead = result.scalars().first()
    if not lead:
        raise HTTPException(status_code=404, detail="Pipeline lead not found")
    if lead.stage != "raw":
        raise HTTPException(status_code=400, detail="Only raw-stage leads accept call results")

    tenant = await tenant_name(request, db)
    settings_row = await pipeline_service.get_settings(db)

    if payload.success:
        pipeline_service.apply_call_success(lead, payload.dict())
        await pipeline_service.write_audit(db, tenant, "AI Calling Agent",
                                           f"External AI call result recorded for {lead.id}: {lead.ai_outcome}.")
    else:
        pipeline_service.apply_call_failure(lead, settings_row.ai_retry_limit,
                                            payload.outcome or "Reported failed by dialer")
        await pipeline_service.write_audit(db, tenant, "AI Calling Agent",
                                           f"External AI call failure recorded for {lead.id}.", status="Failed")
    await db.commit()
    return lead


@router.post("/ai/run-cycle")
async def ai_run_cycle(request: Request, db: AsyncSession = Depends(get_db), user=Depends(get_current_user)):
    """Manually triggers one AI calling sweep for the active workspace (demo/testing helper)."""
    tenant_id = getattr(request.state, "tenant_id", "public")
    processed = await ai_agent.run_cycle_for_tenant(tenant_id, await tenant_name(request, db))
    return {"processed": processed}
